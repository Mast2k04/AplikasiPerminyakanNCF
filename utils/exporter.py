import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def convert_df_to_excel(df, nama_studi_kasus=" "):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluasi Keekonomian"
    ws.views.sheetView[0].showGridLines = True
    
    NAVY_HEADER = "1F4E78"
    WHITE = "FFFFFF"
    ZEBRA_FILL = "F9FAFB"
    MINT_TOTAL = "E2EFDA"
    BORDER_GRAY = "D9D9D9"
    
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1F4E78")
    font_header = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
    font_body = Font(name="Segoe UI", size=10)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    
    fill_header = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
    fill_mint = PatternFill(start_color=MINT_TOTAL, end_color=MINT_TOTAL, fill_type="solid")
    
    thin_border = Border(left=Side(style='thin', color=BORDER_GRAY), right=Side(style='thin', color=BORDER_GRAY), top=Side(style='thin', color=BORDER_GRAY), bottom=Side(style='thin', color=BORDER_GRAY))
    double_bottom = Border(left=Side(style='thin', color=BORDER_GRAY), right=Side(style='thin', color=BORDER_GRAY), top=Side(style='thin', color=BORDER_GRAY), bottom=Side(style='double', color="000000"))

    # Menulis Judul Utama & Nama Studi Kasus dari Input User
    ws['A1'] = "PERHITUNGAN KEEKONOMIAN LAPANGAN MIGAS"
    ws['A1'].font = font_title
    ws['A2'] = f"Nama Lapangan / Studi Kasus: {nama_studi_kasus}"
    ws['A2'].font = Font(name="Segoe UI", size=11, italic=True, bold=True, color="000000")
    
    headers = ["Tahun"] + list(df.columns)
    row_header = 4
    ws.row_dimensions[row_header].height = 28
    
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=row_header, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    current_row = 5
    for index_tahun, row_data in df.iterrows():
        ws.row_dimensions[current_row].height = 20
        cell_tahun = ws.cell(row=current_row, column=1, value=index_tahun)
        cell_tahun.font = font_bold
        cell_tahun.alignment = Alignment(horizontal="center", vertical="center")
        cell_tahun.border = thin_border
        
        is_zebra = (current_row % 2 == 1)
        if is_zebra: cell_tahun.fill = fill_zebra
            
        for col_idx, value in enumerate(row_data, 2):
            cell_val = ws.cell(row=current_row, column=col_idx, value=value)
            cell_val.font = font_body
            cell_val.border = thin_border
            cell_val.alignment = Alignment(horizontal="right", vertical="center")
            
            if is_zebra: cell_val.fill = fill_zebra
                
            nama_kolom = headers[col_idx - 1]
            if "Produksi" in nama_kolom:
                cell_val.number_format = '#,##0.00'
            elif "Tahun" not in nama_kolom:
                cell_val.number_format = '$#,##0.00'
                
            if "NCF" in nama_kolom:
                cell_val.fill = fill_mint
                cell_val.font = font_bold
                
        current_row += 1

    # TAMBAHAN OTOMATIS: Baris TOTAL di bagian bawah lembar Excel
    ws.row_dimensions[current_row].height = 22
    cell_tot_lbl = ws.cell(row=current_row, column=1, value="Total")
    cell_tot_lbl.font = font_bold
    cell_tot_lbl.border = double_bottom
    cell_tot_lbl.alignment = Alignment(horizontal="center", vertical="center")
    
    for col_idx in range(2, len(headers) + 1):
        nama_kolom = headers[col_idx - 1]
        cell_tot = ws.cell(row=current_row, column=col_idx)
        cell_tot.font = font_bold
        cell_tot.border = double_bottom
        
        # Logika penjumlahan akuntansi yang benar
        if "NCF Undiscounted" in nama_kolom:
            cell_tot.value = "" # Kosong karena running total akhir tidak boleh dijumlahkan
        elif "Tahun" not in nama_kolom:
            col_letter = get_column_letter(col_idx)
            # Menghitung sum dari baris ke-5 sampai baris sebelum TOTAL (hanya Tahun 1-10)
            cell_tot.value = f"=SUM({col_letter}6:{col_letter}{current_row-1})"
            cell_tot.number_format = '$#,##0.00'
        elif "Produksi" in nama_kolom:
            col_letter = get_column_letter(col_idx)
            cell_tot.value = f"=SUM({col_letter}6:{col_letter}{current_row-1})"
            cell_tot.number_format = '#,##0.00'

    # Auto-fit kolom
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 4: continue
            if cell.value: max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb.save(output)
    return output.getvalue()